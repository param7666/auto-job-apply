"""
╔══════════════════════════════════════════════════════════╗
║   🤖 AI PC AGENT — FINAL VERSION v7.1                  ║
║   Developer: Parmeshwar Gurlewad                        ║
║   Features: Naukri Auto Apply, LinkedIn Auto Apply,     ║
║             Excel Export, Human Q&A, Tab Management     ║
║   Fix: LinkedIn JS click (bypasses disabled state)      ║
╚══════════════════════════════════════════════════════════╝
"""

import re
import json
import time
import os
from datetime import datetime
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════
# YOUR PROFILE — edit these values
# ══════════════════════════════════════════════════════════
MY_PROFILE = {
    "name":             "Parmeshwar Gurlewad",
    "first_name":       "Parmeshwar",
    "last_name":        "Gurlewad",
    "email":            "parmeshwarg08@gmail.com",
    "phone":            "7666845301",
    "location":         "Hyderabad",
    "gender":           "Male",
    "current_company":  "Pranakshit IT Solution",
    "current_role":     "Software Engineer",
    "experience_years": "1",
    "skills":           "Java, Spring Boot, Spring Framework, Hibernate, Microservices, REST APIs, MySQL, PostgreSQL, React, JavaScript, Git, Maven",
    "resume_path":      r"C:\Users\parme\Desktop\Param_SoftwareEngineer.pdf",
    "expected_ctc":     "400000",
    "current_ctc":      "0",
    "notice_period":    "0",
    "degree":           "Bachelor of Computer Science",
    "college":          "MGM College of Computer Science",
    "passing_year":     "2024",
    "percentage":       "75",
}

AGENT_PROFILE = os.path.expanduser("~") + r"\AI_Agent_Chrome_Profile"
EXCEL_DIR     = os.path.expanduser("~") + r"\Desktop"

# ══════════════════════════════════════════════════════════
# INTENT DETECTION — pure keyword matching, no AI needed
# ══════════════════════════════════════════════════════════

JOB_KEYWORDS = [
    "java full stack developer", "java full stack",
    "java developer", "spring boot developer",
    "python developer", "react developer", "angular developer",
    "full stack developer", "backend developer", "frontend developer",
    "devops engineer", "data engineer", "data scientist",
    "machine learning engineer", "ml engineer",
    "nodejs developer", "software engineer", "software developer",
    "dotnet developer", ".net developer", "php developer",
    "node js", "full stack", "spring boot",
    "java", "python", "react", "angular",
    "backend", "frontend", "devops",
]

KNOWN_CITIES = [
    "hyderabad", "bangalore", "bengaluru", "mumbai", "pune",
    "chennai", "delhi", "gurgaon", "noida", "kolkata",
    "ahmedabad", "jaipur", "coimbatore", "kochi", "indore",
    "bhopal", "surat", "lucknow", "chandigarh", "nagpur",
    "vizag", "visakhapatnam", "remote",
]

def parse_command(command):
    c = command.lower().strip()

    on_linkedin = any(w in c for w in ["linkedin", "linked in", "linkdin", "linkd in", "linked-in", "linkedn"])
    on_naukri   = any(w in c for w in ["naukri", "nokri", "noukri", "naukari", "naukrii"])
    on_youtube  = ("youtube" in c or "you tube" in c or bool(re.search(r"\byt\b", c)))
    on_google   = "google" in c

    want_apply  = any(w in c for w in ["apply", "auto apply", "submit", "send application"])
    want_search = any(w in c for w in ["search", "find", "show", "get", "list", "save", "open", "go", "browse"])

    job_title = "Java Developer"
    for kw in JOB_KEYWORDS:
        if kw in c:
            job_title = kw.title()
            break

    location = ""
    for city in KNOWN_CITIES:
        if city in c:
            location = city.title()
            break

    max_apply = 5
    nums = re.findall(r"\b(\d+)\b", c)
    if nums:
        max_apply = int(nums[0])

    if on_youtube:
        query = re.sub(r"\b(youtube|search|find|on)\b", "", c).strip()
        return {"intent": "youtube_search", "query": query, "job_title": job_title, "location": location, "max_apply": max_apply}

    if on_google:
        query = re.sub(r"\b(google|search|find|on)\b", "", c).strip()
        return {"intent": "google_search", "query": query, "job_title": job_title, "location": location, "max_apply": max_apply}

    if on_linkedin:
        intent = "linkedin_apply" if want_apply else "linkedin_search"
        return {"intent": intent, "job_title": job_title, "location": location, "max_apply": max_apply, "query": ""}

    if on_naukri or want_apply:
        intent = "naukri_search" if (want_search and not want_apply) else "naukri_auto_apply"
        return {"intent": intent, "job_title": job_title, "location": location, "max_apply": max_apply, "query": ""}

    return {"intent": "unknown"}

# ══════════════════════════════════════════════════════════
# SMART ANSWER SYSTEM
# ══════════════════════════════════════════════════════════

QUICK_ANSWERS = {
    "notice period":          MY_PROFILE["notice_period"],
    "notice":                 MY_PROFILE["notice_period"],
    "current ctc":            MY_PROFILE["current_ctc"],
    "current salary":         MY_PROFILE["current_ctc"],
    "expected ctc":           MY_PROFILE["expected_ctc"],
    "expected salary":        MY_PROFILE["expected_ctc"],
    "experience":             MY_PROFILE["experience_years"],
    "years of experience":    MY_PROFILE["experience_years"],
    "how many years":         MY_PROFILE["experience_years"],
    "total experience":       MY_PROFILE["experience_years"],
    "first name":             MY_PROFILE["first_name"],
    "last name":              MY_PROFILE["last_name"],
    "full name":              MY_PROFILE["name"],
    "email":                  MY_PROFILE["email"],
    "phone":                  MY_PROFILE["phone"],
    "mobile":                 MY_PROFILE["phone"],
    "contact":                MY_PROFILE["phone"],
    "location":               MY_PROFILE["location"],
    "city":                   MY_PROFILE["location"],
    "gender":                 MY_PROFILE["gender"],
    "degree":                 MY_PROFILE["degree"],
    "college":                MY_PROFILE["college"],
    "university":             MY_PROFILE["college"],
    "passing year":           MY_PROFILE["passing_year"],
    "graduation year":        MY_PROFILE["passing_year"],
    "percentage":             MY_PROFILE["percentage"],
    "cgpa":                   MY_PROFILE["percentage"],
    "skills":                 MY_PROFILE["skills"],
    "relocate":               "Yes",
    "willing to relocate":    "Yes",
    "authorized":             "Yes",
    "work in india":          "Yes",
    "immediate":              "Yes",
    "project":                "Yes",
    "worked on":              "Yes",
    "fresher or experienced": "Experienced",
    "fresher":                "Experienced",
    "currently employed":     "Yes",
    "currently working":      "Yes",
    "java":                   "Yes",
    "spring boot":            "Yes",
    "microservices":          "Yes",
}

UI_SKIP_WORDS = [
    "enter keyword", "keyword", "designation", "companies",
    "enter skills", "search jobs", "search", "where",
    "city, state", "enter location", "what", "job title",
    "enter city", "find jobs",
]

def auto_answer(question_text, options=None):
    q = question_text.lower().strip()
    if not q or len(q) < 3:
        return None
    if any(skip in q for skip in UI_SKIP_WORDS):
        return None
    for keyword, answer in QUICK_ANSWERS.items():
        if keyword in q:
            return str(answer)
    if options:
        for keyword, answer in QUICK_ANSWERS.items():
            if keyword in q:
                ans_lower = str(answer).lower()
                for opt in options:
                    if ans_lower in opt.lower() or opt.lower() in ans_lower:
                        return opt
    return None

def ask_user(question_text, options=None):
    print("\n" + "─" * 56)
    print(f"  ❓ BOT QUESTION:")
    print(f"  {question_text}")
    if options:
        print()
        for i, opt in enumerate(options, 1):
            print(f"    [{i}] {opt}")
        print(f"\n  Type number OR your own answer:")
    else:
        print("  Type your answer:")
    answer = input("  👉 ").strip()
    if options and answer.isdigit():
        idx = int(answer) - 1
        if 0 <= idx < len(options):
            answer = options[idx]
            print(f"  ✅ Selected: '{answer}'")
    print("─" * 56)
    return answer if answer else "Yes"

def smart_answer(question_text, options=None):
    answer = auto_answer(question_text, options)
    if answer:
        print(f"      ⚡ Auto: '{question_text[:35]}' → '{answer}'")
        return answer
    return ask_user(question_text, options)

# ══════════════════════════════════════════════════════════
# BROWSER UTILITIES
# ══════════════════════════════════════════════════════════

def safe_goto(page, url, timeout=40000):
    try:
        page.goto(url, wait_until="domcontentloaded", timeout=timeout)
        time.sleep(2)
        return True
    except PlaywrightTimeout:
        time.sleep(2)
        return True
    except Exception as e:
        print(f"   ❌ Nav error: {e}")
        return False

def fill_field(page, selectors, value):
    for sel in selectors:
        try:
            el = page.query_selector(sel)
            if el and el.is_visible():
                el.click(); time.sleep(0.2)
                el.triple_click()
                el.fill(str(value))
                return True
        except:
            continue
    return False

def close_extra_tabs(browser, keep_page):
    try:
        for p in browser.pages:
            if p != keep_page:
                try: p.close()
                except: pass
    except:
        pass
    return keep_page

def check_success_naukri(page):
    try:
        url = page.url.lower()
        if "saveapply" in url or "applyconfirmation" in url:
            return True
        title = page.title().lower()
        if "apply confirmation" in title or "applied" in title:
            return True
        body = page.inner_text("body").lower()[:5000]
        return any(t in body for t in [
            "applied successfully", "application submitted",
            "you have already applied", "application has been sent",
            "successfully applied", "thank you for applying",
            "your application has been", 'applied to "',
            "apply confirmation",
        ])
    except:
        return False

def is_external_apply(page):
    naukri_apply_selectors = [
        'button[id="apply-button"]',
        '#apply-button',
        'button.apply-button',
    ]
    for sel in naukri_apply_selectors:
        try:
            btn = page.query_selector(sel)
            if btn and btn.is_visible():
                txt = (btn.inner_text() or "").lower()
                if "company site" not in txt and "employer site" not in txt:
                    return False
        except:
            pass

    external_texts = [
        "apply on company site", "apply on employer site",
        "apply at company", "apply on company's site",
    ]
    for sel in [
        'a[href]:has-text("Apply on company")',
        'button:has-text("Apply on company")',
        'a[href]:has-text("Apply on employer")',
        'button:has-text("Apply on employer")',
        'a[href]:has-text("Apply at company")',
    ]:
        try:
            el = page.query_selector(sel)
            if el and el.is_visible():
                txt = el.inner_text().lower()
                if any(t in txt for t in external_texts):
                    return True
        except:
            pass

    return False

# ══════════════════════════════════════════════════════════
# NAUKRI — CHATBOT MODAL HANDLER
# ══════════════════════════════════════════════════════════

def handle_naukri_chatbot(page):
    print("      💬 Chatbot detected — answering questions...")

    for round_num in range(20):
        time.sleep(1.5)

        if check_success_naukri(page):
            return True

        question_text = ""
        for sel in [
            '.chatbot-message', '[class*="question"] p',
            '.ssrc__question', '[class*="chatbot"] p',
            '[class*="questionnaire"] label', '.modal p',
        ]:
            try:
                el = page.query_selector(sel)
                if el and el.is_visible():
                    txt = el.inner_text().strip()
                    if txt and len(txt) > 3:
                        question_text = txt
                        break
            except:
                pass

        if question_text:
            print(f"      ❓ {question_text[:70]}")

        radio_done = False
        try:
            radios = page.query_selector_all('input[type="radio"]')
            visible_radios = [r for r in radios if r.is_visible()]

            if visible_radios:
                option_map = []
                for radio in visible_radios:
                    try:
                        rid = radio.get_attribute("id") or ""
                        val = radio.get_attribute("value") or ""
                        label_text = val
                        if rid:
                            lbl = page.query_selector(f'label[for="{rid}"]')
                            if lbl and lbl.is_visible():
                                label_text = lbl.inner_text().strip()
                        if label_text:
                            option_map.append((label_text, radio))
                    except:
                        continue

                if option_map:
                    labels = [o[0] for o in option_map]
                    print(f"      🔘 Options: {labels}")

                    answer = auto_answer(question_text or "", labels)

                    if not answer:
                        yes_options = [o for o in option_map if "yes" in o[0].lower()]
                        if yes_options and len(option_map) == 2:
                            answer = yes_options[0][0]
                            print(f"      ⚡ Auto-Yes: '{answer}'")
                        else:
                            answer = ask_user(question_text or "Select best option", labels)

                    clicked = False
                    answer_lower = answer.lower()
                    for label, radio_el in option_map:
                        if answer_lower in label.lower() or label.lower() in answer_lower:
                            try:
                                radio_el.click(); clicked = True; radio_done = True
                                print(f"      ✅ Selected: '{label}'")
                                time.sleep(0.5); break
                            except:
                                pass

                    if not clicked:
                        for label, radio_el in option_map:
                            if "yes" in label.lower():
                                try: radio_el.click(); clicked = True; radio_done = True; break
                                except: pass
                    if not clicked:
                        try: option_map[0][1].click(); radio_done = True
                        except: pass
        except:
            pass

        try:
            inputs = page.query_selector_all('input[type="text"], input[type="number"], input[type="tel"]')
            for inp in inputs:
                if not inp.is_visible(): continue
                current = inp.input_value() or ""
                if current.strip(): continue
                placeholder = (inp.get_attribute("placeholder") or "").strip()
                if any(skip in placeholder.lower() for skip in UI_SKIP_WORDS):
                    continue
                label_text = question_text or placeholder
                try:
                    inp_id = inp.get_attribute("id") or ""
                    if inp_id:
                        lbl = page.query_selector(f'label[for="{inp_id}"]')
                        if lbl: label_text = lbl.inner_text().strip()
                except: pass
                if not label_text: continue
                answer = smart_answer(label_text)
                if answer:
                    inp.triple_click(); inp.fill(answer); time.sleep(0.3)
        except:
            pass

        try:
            for ta in page.query_selector_all("textarea"):
                if ta.is_visible() and not (ta.input_value() or "").strip():
                    ta.fill("Yes, I have worked on multiple Java projects including microservices and REST APIs using Spring Boot.")
        except:
            pass

        try:
            for sel_el in page.query_selector_all("select"):
                if not sel_el.is_visible(): continue
                answer = smart_answer(question_text or "Select option")
                try: sel_el.select_option(label=answer)
                except:
                    try: sel_el.select_option(index=1)
                    except: pass
        except:
            pass

        time.sleep(0.5)
        btn_clicked = False
        for btn_text in ["Save", "Next", "Submit", "Continue", "Proceed", "Done"]:
            for sel in [
                f'button:has-text("{btn_text}")',
                f'[class*="btn"]:has-text("{btn_text}")',
            ]:
                try:
                    btn = page.query_selector(sel)
                    if btn and btn.is_visible():
                        btn.click()
                        print(f"      ▶ Clicked '{btn_text}'")
                        btn_clicked = True
                        time.sleep(1.5)
                        break
                except:
                    pass
            if btn_clicked:
                break

        if check_success_naukri(page):
            return True

        if not btn_clicked and not radio_done:
            try: page.keyboard.press("Enter"); time.sleep(1)
            except: pass
            break

    return check_success_naukri(page)

# ══════════════════════════════════════════════════════════
# NAUKRI — APPLY FLOW
# ══════════════════════════════════════════════════════════

def naukri_apply_flow(page):
    for step in range(15):
        time.sleep(1.5)

        if check_success_naukri(page):
            return True

        chatbot_open = False
        for sel in ['.chatbot-container', '[class*="chatbot"]', '[class*="questionnaire"]', '.apply-questionnaire']:
            try:
                el = page.query_selector(sel)
                if el and el.is_visible():
                    chatbot_open = True; break
            except: pass

        try:
            if any(r.is_visible() for r in page.query_selector_all('input[type="radio"]')):
                chatbot_open = True
        except: pass

        if chatbot_open:
            return handle_naukri_chatbot(page)

        fill_field(page, ['input[placeholder*="First name"]', 'input[id*="firstName"]'], MY_PROFILE["first_name"])
        fill_field(page, ['input[placeholder*="Last name"]', 'input[id*="lastName"]'], MY_PROFILE["last_name"])
        fill_field(page, ['input[type="email"]'], MY_PROFILE["email"])
        fill_field(page, ['input[type="tel"]', 'input[placeholder*="phone"]'], MY_PROFILE["phone"])
        fill_field(page, ['input[placeholder*="current ctc"]', 'input[id*="currentCTC"]'], MY_PROFILE["current_ctc"])
        fill_field(page, ['input[placeholder*="expected"]', 'input[id*="expectedCTC"]'], MY_PROFILE["expected_ctc"])
        fill_field(page, ['input[placeholder*="notice"]', 'input[id*="notice"]'], MY_PROFILE["notice_period"])
        fill_field(page, ['input[placeholder*="experience"]', 'input[id*="experience"]'], MY_PROFILE["experience_years"])

        try:
            for fi in page.query_selector_all('input[type="file"]'):
                try: fi.set_input_files(MY_PROFILE["resume_path"]); time.sleep(1); break
                except: pass
        except: pass

        submitted = False
        for sel in ['button:has-text("Apply Now")', 'button:has-text("Apply")',
                    'button:has-text("Submit")', 'button[type="submit"]']:
            try:
                btn = page.query_selector(sel)
                if btn and btn.is_visible():
                    btn.click(); time.sleep(2); submitted = True; break
            except: pass

        if check_success_naukri(page):
            return True
        if not submitted:
            break

    return check_success_naukri(page)

# ══════════════════════════════════════════════════════════
# NAUKRI — SCRAPE JOBS
# ══════════════════════════════════════════════════════════

def scrape_naukri_jobs(page):
    print("   🕷️  Scraping Naukri jobs...")
    for _ in range(4):
        page.mouse.wheel(0, 800); time.sleep(0.7)
    try:
        page.wait_for_selector(".srp-jobtuple-wrapper, .jobTuple", timeout=10000)
    except: pass

    cards = []
    for sel in [".srp-jobtuple-wrapper", ".jobTuple", "article.jobTuple"]:
        c = page.query_selector_all(sel)
        if c: cards = c; break

    print(f"   📦 Found {len(cards)} job cards")
    jobs = []
    for card in cards:
        try:
            t  = card.query_selector("a.title, .title a, h2 a")
            co = card.query_selector("a.comp-name, .comp-name")
            ex = card.query_selector(".expwdth, .exp-wrap li")
            sa = card.query_selector(".sal-wrap li, .salary")
            lo = card.query_selector(".loc-wrap li, .location")
            po = card.query_selector(".job-post-day, .postDays, span.type-time")
            title = t.inner_text().strip() if t else "N/A"
            href  = (t.get_attribute("href") or "") if t else ""
            link  = href if href.startswith("http") else "https://www.naukri.com" + href
            if title != "N/A":
                jobs.append({
                    "title":    title,
                    "company":  co.inner_text().strip() if co else "N/A",
                    "exp":      ex.inner_text().strip() if ex else "N/A",
                    "salary":   sa.inner_text().strip() if sa else "Not disclosed",
                    "location": lo.inner_text().strip() if lo else "N/A",
                    "posted":   po.inner_text().strip() if po else "N/A",
                    "link":     link,
                    "applied":  "Pending",
                })
        except: continue

    seen_links = set()
    seen_company_title = {}
    deduped = []
    for job in jobs:
        link    = job.get("link", "")
        company = job.get("company", "").lower().strip()
        if link and link in seen_links:
            continue
        company_count = seen_company_title.get(company, 0)
        if company_count >= 3:
            continue
        seen_links.add(link)
        seen_company_title[company] = company_count + 1
        deduped.append(job)

    removed = len(jobs) - len(deduped)
    if removed > 0:
        print(f"   🔁 Removed {removed} duplicate/spam listings")
    print(f"   ✅ Scraped {len(deduped)} unique jobs")
    return deduped

# ══════════════════════════════════════════════════════════
# NAUKRI — AUTO APPLY
# ══════════════════════════════════════════════════════════

def naukri_auto_apply(browser, page, job_title, location="", max_apply=5):
    print(f"\n{'='*58}")
    print(f"  🤖 NAUKRI AUTO APPLY: '{job_title}'")
    print(f"  ⏰ Last 24h filter ON  |  🎯 Max: {max_apply}")
    print(f"{'='*58}\n")

    safe_goto(page, "https://www.naukri.com/mnjuser/homepage", timeout=30000)
    time.sleep(2)
    if "login" in page.url or "mnjuser" not in page.url:
        print("⚠️  Not logged in! Please login to Naukri in the browser, then press ENTER here...")
        input()

    loc_part = f"-in-{location.lower().replace(' ', '-')}" if location and location.lower() not in ["any", ""] else ""
    keyword  = job_title.lower().replace(" ", "-")

    search_attempts = []
    if loc_part:
        search_attempts += [
            ("&freshness=1", loc_part, "24h + " + location),
            ("&freshness=7", loc_part, "7 days + " + location),
            ("",             loc_part, "all time + " + location),
            ("&freshness=1", "",       "24h (any location)"),
            ("",             "",       "all time (any location)"),
        ]
    else:
        search_attempts += [
            ("&freshness=1", "", "24 hours"),
            ("&freshness=7", "", "7 days"),
            ("",             "", "all time"),
        ]

    jobs = []
    for fresh_param, lp, label in search_attempts:
        url = f"https://www.naukri.com/{keyword}-jobs{lp}?experienceMin=0&experienceMax=2{fresh_param}"
        print(f"   🔗 Trying: [{label}]")
        safe_goto(page, url, timeout=40000)
        time.sleep(2)
        jobs = scrape_naukri_jobs(page)
        jobs = [j for j in jobs if j.get("title") and j.get("company")]
        if len(jobs) >= max_apply:
            print(f"   ✅ Found {len(jobs)} jobs [{label}]")
            break
        elif jobs:
            print(f"   ⚠️  Only {len(jobs)} jobs [{label}] — trying wider...")
        else:
            print(f"   ⚠️  No jobs [{label}] — trying wider...")

    if not jobs:
        print("   ❌ No jobs found. Try a different job title.")
        save_naukri_excel([], job_title, 0)
        return []

    applied_count = skipped_ext = skipped_err = 0
    applied_jobs  = []
    search_url    = page.url
    already_applied = getattr(naukri_auto_apply, "_applied_links", set())

    for job in jobs:
        if job.get("link") in already_applied:
            print(f"  ⏭️  Already applied this session — skipping {job['title']}")
            job["applied"] = "Skipped (Already applied)"
            skipped_err += 1
            continue
        if applied_count >= max_apply:
            print(f"\n✅ Reached limit of {max_apply}!")
            break
        if not job.get("link"):
            continue

        print(f"\n  [{applied_count+1}/{max_apply}] {job['title']} @ {job['company']}")
        print(f"  💰 {job['salary']}  📍 {job['location']}  ⏰ {job['posted']}")

        tabs_before = len(browser.pages)
        safe_goto(page, job["link"], timeout=30000)

        try:
            page.wait_for_selector(
                '#apply-button, button[id="apply-button"], .apply-button, a:has-text("Apply on company")',
                timeout=8000
            )
        except:
            pass
        time.sleep(1)

        apply_btn = None
        is_ext    = False

        for sel in ['button[id="apply-button"]', '#apply-button']:
            try:
                btn = page.query_selector(sel)
                if btn and btn.is_visible():
                    txt = (btn.inner_text() or "").lower().strip()
                    print(f"     🔍 Found apply btn: '{txt[:40]}'")
                    if "company site" in txt or "employer site" in txt:
                        is_ext = True
                    else:
                        apply_btn = btn
                    break
            except: pass

        if not apply_btn and not is_ext:
            for sel in [
                'a[href]:has-text("Apply on company site")',
                'a[href]:has-text("Apply on employer site")',
                'button:has-text("Apply on company site")',
                'button:has-text("Apply on employer site")',
            ]:
                try:
                    btn = page.query_selector(sel)
                    if btn and btn.is_visible():
                        print(f"     🔍 External link found: '{(btn.inner_text() or '')[:40]}'")
                        is_ext = True
                        break
                except: pass

        if not apply_btn and not is_ext:
            for sel in ['.apply-button', '.btn-apply', '[class*="apply-button"]']:
                try:
                    btn = page.query_selector(sel)
                    if btn and btn.is_visible():
                        txt = (btn.inner_text() or "").lower().strip()
                        print(f"     🔍 Class apply btn: '{txt[:40]}'")
                        if "company site" not in txt and "employer site" not in txt:
                            apply_btn = btn
                        break
                except: pass

        if is_ext:
            print("     🚫 External apply — skipping")
            job["applied"] = "Skipped (External)"; skipped_ext += 1
            safe_goto(page, search_url, timeout=30000); time.sleep(1); continue

        if not apply_btn:
            print("     ⏭️  No Apply button found — skipping")
            job["applied"] = "Skipped (No button)"; skipped_err += 1
            safe_goto(page, search_url, timeout=30000); time.sleep(1); continue

        apply_btn.click()
        time.sleep(2.5)

        tabs_after = len(browser.pages)
        active_page = page
        if tabs_after > tabs_before:
            print(f"     🆕 New tab opened — switching...")
            for p in browser.pages:
                if p != page:
                    active_page = p
                    active_page.bring_to_front()
                    time.sleep(1.5)
                    print(f"     📄 {active_page.url[:65]}")
                    break

        if check_success_naukri(active_page):
            print("     🎉 Already applied! (Confirmation page)")
            applied_count += 1
            job["applied"] = "Applied ✅"
            applied_jobs.append({**job, "time": datetime.now().strftime("%d %b %Y %H:%M")})
            close_extra_tabs(browser, page)
            safe_goto(page, search_url, timeout=30000); time.sleep(1.5); continue

        if is_external_apply(active_page):
            print("     🚫 Redirected to external — skipping")
            job["applied"] = "Skipped (External)"; skipped_ext += 1
            close_extra_tabs(browser, page)
            safe_goto(page, search_url, timeout=30000); time.sleep(1); continue

        print("     📝 Filling application...")
        success = naukri_apply_flow(active_page)

        if success:
            applied_count += 1
            job["applied"] = "Applied ✅"
            applied_jobs.append({**job, "time": datetime.now().strftime("%d %b %Y %H:%M")})
            if not hasattr(naukri_auto_apply, "_applied_links"):
                naukri_auto_apply._applied_links = set()
            naukri_auto_apply._applied_links.add(job.get("link", ""))
            print(f"     🎉 APPLIED! Total: {applied_count}")
        else:
            job["applied"] = "Skipped (Form error)"; skipped_err += 1
            print("     ⚠️  Could not complete")
            try: active_page.keyboard.press("Escape")
            except: pass

        close_extra_tabs(browser, page)
        print("     🔄 Returning to search...")
        safe_goto(page, search_url, timeout=30000)
        time.sleep(1.5)

    print(f"\n{'='*58}")
    print(f"  🏆 Applied: {applied_count}  |  External skipped: {skipped_ext}  |  Errors: {skipped_err}")
    print(f"{'='*58}")
    # Always save Excel — even if 0 applied
    save_naukri_excel(jobs, job_title, applied_count)
    return applied_jobs


def naukri_search_only(page, job_title, location=""):
    print(f"\n🔍 Naukri Search: '{job_title}' (last 24h)")
    safe_goto(page, "https://www.naukri.com/mnjuser/homepage", timeout=30000)
    loc_part = f"-in-{location.lower()}" if location and location.lower() not in ["any", ""] else ""
    keyword  = job_title.lower().replace(" ", "-")
    url = f"https://www.naukri.com/{keyword}-jobs{loc_part}?freshness=1"
    safe_goto(page, url, timeout=40000); time.sleep(2)
    jobs = scrape_naukri_jobs(page)
    # Always save Excel
    save_naukri_excel(jobs, job_title, 0)
    if jobs:
        print(f"✅ Saved {len(jobs)} jobs to Excel!")
    else:
        print("❌ No jobs found.")

# ══════════════════════════════════════════════════════════
# LINKEDIN — AUTO APPLY  (v7.1 — JS click fix)
# ══════════════════════════════════════════════════════════

def linkedin_auto_apply(browser, page, job_title, location="", max_apply=5):
    print(f"\n{'='*58}")
    print(f"  💼 LINKEDIN AUTO APPLY: '{job_title}'")
    print(f"  🎯 Max: {max_apply} | Easy Apply only")
    print(f"{'='*58}\n")

    safe_goto(page, "https://www.linkedin.com/feed/", timeout=30000)
    time.sleep(2)
    if "login" in page.url or "authwall" in page.url:
        print("⚠️  Not logged in! Please login to LinkedIn in the browser, then press ENTER...")
        input()

    kw  = job_title.replace(" ", "%20")
    loc = location.replace(" ", "%20") if location and location.lower() not in ["any", ""] else "India"
    url = f"https://www.linkedin.com/jobs/search/?keywords={kw}&location={loc}&f_LF=f_AL&f_TPR=r86400"
    print(f"   🔗 {url}")
    safe_goto(page, url, timeout=40000)
    time.sleep(3)

    applied_count = skipped = 0
    jobs_data = []

    # Scroll to load more job cards
    for _ in range(3):
        page.mouse.wheel(0, 1000); time.sleep(0.8)

    try:
        page.wait_for_selector(
            ".jobs-search-results__list-item, .scaffold-layout__list-item",
            timeout=10000
        )
    except: pass

    cards = page.query_selector_all(
        ".jobs-search-results__list-item, .scaffold-layout__list-item"
    )
    print(f"   📦 Found {len(cards)} job cards\n")

    for idx in range(len(cards)):
        if applied_count >= max_apply:
            print(f"\n✅ Reached limit of {max_apply}!")
            break

        try:
            # Re-fetch fresh every iteration — avoids stale DOM
            cards_fresh = page.query_selector_all(
                ".jobs-search-results__list-item, .scaffold-layout__list-item"
            )
            if idx >= len(cards_fresh):
                break

            card = cards_fresh[idx]

            # ── FIX: JS click bypasses LinkedIn's disabled-attribute check ──
            page.evaluate("(el) => el.scrollIntoView({block:'center'})", card)
            time.sleep(0.3)
            page.evaluate("(el) => el.click()", card)
            time.sleep(3)  # wait for detail panel to fully load

            # Get job details from right panel
            title_el = page.query_selector(
                "h1.t-24, "
                ".job-details-jobs-unified-top-card__job-title h1, "
                "h1.jobs-unified-top-card__job-title"
            )
            company_el = page.query_selector(
                ".job-details-jobs-unified-top-card__company-name a, "
                ".job-details-jobs-unified-top-card__subtitle-top-block a, "
                ".jobs-unified-top-card__company-name a"
            )
            title   = title_el.inner_text().strip()   if title_el   else f"Job #{idx+1}"
            company = company_el.inner_text().strip()  if company_el else "N/A"

            print(f"  [{applied_count+1}/{max_apply}] {title} @ {company}")

            # ── Find Easy Apply button ────────────────────────────────────
            easy_apply_btn = None
            easy_apply_selectors = [
                'button.jobs-apply-button:has-text("Easy Apply")',
                'button[aria-label*="Easy Apply"]',
                'button:has-text("Easy Apply")',
                '.jobs-apply-button:has-text("Easy Apply")',
            ]

            # Wait up to 5s for Easy Apply button to appear
            for sel in easy_apply_selectors:
                try:
                    page.wait_for_selector(sel, timeout=4000)
                    break
                except: pass

            for sel in easy_apply_selectors:
                try:
                    btn = page.query_selector(sel)
                    if btn and btn.is_visible():
                        easy_apply_btn = btn
                        break
                except: pass

            if not easy_apply_btn:
                print("     ⏭️  No Easy Apply — skipping")
                skipped += 1
                jobs_data.append({
                    "title": title, "company": company,
                    "applied": "Skipped (No Easy Apply)",
                    "time": datetime.now().strftime("%d %b %Y %H:%M"),
                })
                continue

            # ── FIX: JS click on Easy Apply button ───────────────────────
            print("     🖱️  Clicking Easy Apply (JS)…")
            try:
                page.evaluate("(el) => el.click()", easy_apply_btn)
            except:
                try:
                    easy_apply_btn.click(force=True, timeout=5000)
                except Exception as ce:
                    print(f"     ⚠️  Click failed: {ce}")
                    skipped += 1
                    jobs_data.append({
                        "title": title, "company": company,
                        "applied": "Skipped (Click failed)",
                        "time": datetime.now().strftime("%d %b %Y %H:%M"),
                    })
                    continue

            time.sleep(2)

            # Check modal opened
            modal_open = False
            for sel in ['.jobs-easy-apply-modal', '[data-test-modal]', 'div[role="dialog"]']:
                try:
                    el = page.query_selector(sel)
                    if el and el.is_visible():
                        modal_open = True; break
                except: pass

            if not modal_open:
                print("     ⏭️  Modal didn't open — skipping")
                skipped += 1
                jobs_data.append({
                    "title": title, "company": company,
                    "applied": "Skipped (No modal)",
                    "time": datetime.now().strftime("%d %b %Y %H:%M"),
                })
                continue

            print("     📝 Filling Easy Apply modal…")
            success = linkedin_fill_modal(page)

            if success:
                applied_count += 1
                jobs_data.append({
                    "title": title, "company": company,
                    "applied": "Applied ✅",
                    "time": datetime.now().strftime("%d %b %Y %H:%M"),
                })
                print(f"     🎉 APPLIED! Total: {applied_count}")
            else:
                skipped += 1
                jobs_data.append({
                    "title": title, "company": company,
                    "applied": "Skipped (Form error)",
                    "time": datetime.now().strftime("%d %b %Y %H:%M"),
                })
                print("     ⚠️  Could not complete — discarding")
                try:
                    page.keyboard.press("Escape"); time.sleep(1)
                    discard = page.query_selector('button:has-text("Discard")')
                    if discard and discard.is_visible():
                        discard.click(); time.sleep(1)
                except: pass

        except Exception as e:
            print(f"     ❌ {e}")
            skipped += 1
            jobs_data.append({
                "title": f"Job #{idx+1}", "company": "N/A",
                "applied": "Error",
                "time": datetime.now().strftime("%d %b %Y %H:%M"),
            })
            try:
                page.keyboard.press("Escape"); time.sleep(1)
                discard = page.query_selector('button:has-text("Discard")')
                if discard and discard.is_visible():
                    discard.click(); time.sleep(1)
            except: pass
            continue

    print(f"\n{'='*58}")
    print(f"  🏆 Applied: {applied_count}  |  Skipped: {skipped}")
    print(f"{'='*58}")
    # Always save Excel — even if 0 applied (so dashboard can show the list)
    save_linkedin_excel(jobs_data, job_title, applied_count)
    return jobs_data


def linkedin_fill_modal(page):
    """
    Fill LinkedIn Easy Apply multi-step modal.
    Re-fetches every element fresh each step to avoid stale DOM errors.
    """
    for step in range(15):
        time.sleep(1.5)

        # Check if modal is still open
        modal = None
        for sel in ['.jobs-easy-apply-modal', '[data-test-modal]', 'div[role="dialog"]']:
            try:
                el = page.query_selector(sel)
                if el and el.is_visible():
                    modal = el; break
            except: pass

        if not modal:
            time.sleep(1)
            try:
                body = page.inner_text("body").lower()[:3000]
                if any(t in body for t in ["application submitted", "applied", "your application was sent"]):
                    return True
            except: pass
            return True  # modal gone = submitted

        # Phone number
        fill_field(page, [
            'input[id*="phoneNumber"]',
            'input[placeholder*="phone"]',
            'input[placeholder*="Phone"]',
        ], MY_PROFILE["phone"])

        # Text / Number inputs
        try:
            inputs = page.query_selector_all('input[type="text"], input[type="number"]')
            for inp in inputs:
                if not inp.is_visible(): continue
                if (inp.input_value() or "").strip(): continue
                label_text = ""
                try:
                    inp_id = inp.get_attribute("id") or ""
                    if inp_id:
                        lbl = page.query_selector(f'label[for="{inp_id}"]')
                        if lbl: label_text = lbl.inner_text().strip()
                except: pass
                if not label_text:
                    label_text = inp.get_attribute("placeholder") or "years of experience"
                answer = smart_answer(label_text)
                if answer:
                    inp.triple_click(); inp.fill(answer); time.sleep(0.2)
        except: pass

        # Radio buttons
        try:
            radios = page.query_selector_all('input[type="radio"]')
            visible = [r for r in radios if r.is_visible()]
            if visible:
                option_map = []
                for r in visible:
                    try:
                        rid = r.get_attribute("id") or ""
                        lbl_el = page.query_selector(f'label[for="{rid}"]') if rid else None
                        label = lbl_el.inner_text().strip() if lbl_el else (r.get_attribute("value") or "")
                        if label: option_map.append((label, r))
                    except: pass

                if option_map:
                    labels = [o[0] for o in option_map]
                    q_text = ""
                    try:
                        q_el = page.query_selector('.jobs-easy-apply-form-element legend, .fb-form-element-label')
                        if q_el: q_text = q_el.inner_text().strip()
                    except: pass
                    answer = smart_answer(q_text or "Select best option", labels)
                    ans_lower = answer.lower()
                    clicked = False
                    for label, rel in option_map:
                        if ans_lower in label.lower() or label.lower() in ans_lower:
                            try: rel.click(); clicked = True; time.sleep(0.3); break
                            except: pass
                    if not clicked:
                        for label, rel in option_map:
                            if "yes" in label.lower():
                                try: rel.click(); clicked = True; break
                                except: pass
                    if not clicked and option_map:
                        try: option_map[0][1].click()
                        except: pass
        except: pass

        # Dropdowns
        try:
            for sel_el in page.query_selector_all("select"):
                if not sel_el.is_visible(): continue
                try: sel_el.select_option(index=1)
                except: pass
        except: pass

        # Click button — always re-query fresh, never store reference
        btn_clicked = False
        for btn_text in ["Submit application", "Submit", "Review", "Next", "Continue", "Done"]:
            try:
                btn = page.query_selector(f'button:has-text("{btn_text}")')
                if btn and btn.is_visible() and btn.is_enabled():
                    btn.click()
                    print(f"     ▶ '{btn_text}'")
                    btn_clicked = True
                    time.sleep(2)
                    break
            except: pass

        if not btn_clicked:
            # Try footer buttons inside modal
            try:
                footer_btns = page.query_selector_all(
                    '.jobs-easy-apply-modal footer button, [data-test-modal] footer button'
                )
                for btn in reversed(footer_btns):
                    if btn.is_visible() and btn.is_enabled():
                        txt = (btn.inner_text() or "").strip()
                        if txt:
                            btn.click()
                            print(f"     ▶ Footer: '{txt}'")
                            btn_clicked = True
                            time.sleep(2)
                            break
            except: pass

        if not btn_clicked:
            break  # stuck — give up

        # Check success
        try:
            body = page.inner_text("body").lower()[:3000]
            if any(t in body for t in ["application submitted", "your application was sent", "done"]):
                return True
        except: pass

    return False


def linkedin_search(page, job_title, location=""):
    kw  = job_title.replace(" ", "%20")
    loc = location.replace(" ", "%20") if location and location.lower() not in ["any", ""] else "India"
    url = f"https://www.linkedin.com/jobs/search/?keywords={kw}&location={loc}&f_TPR=r86400"
    print(f"\n🔗 Opening LinkedIn: {job_title} in {loc}")
    safe_goto(page, url, timeout=40000)
    print("✅ LinkedIn jobs opened!")
    print("   💡 Use the dashboard Apply button to auto-apply.")

# ══════════════════════════════════════════════════════════
# EXCEL EXPORT
# ══════════════════════════════════════════════════════════

def _excel_style(ws, title_text, headers, rows, col_widths):
    DARK = "1F3864"; WHITE = "FFFFFF"; GOLD = "FFD700"
    GREEN = "C6EFCE"; YELLOW = "FFEB9C"; RED = "FFC7CE"; LIGHT = "D9E1F2"
    thin = Side(style="thin", color="CCCCCC")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    col_count = len(headers)
    last_col  = get_column_letter(col_count)

    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value = title_text
    c.font  = Font(name="Arial", bold=True, size=13, color=GOLD)
    c.fill  = PatternFill("solid", fgColor=DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    for ci, h in enumerate(headers, 1):
        c = ws.cell(2, ci, h)
        c.font  = Font(name="Arial", bold=True, size=11, color=WHITE)
        c.fill  = PatternFill("solid", fgColor=DARK)
        c.alignment = Alignment(horizontal="center")
        c.border = bdr
    ws.row_dimensions[2].height = 22

    for i, row_data in enumerate(rows, 1):
        r = i + 2
        status = str(row_data[-1])
        if "Applied" in status:    fc = GREEN
        elif "External" in status: fc = YELLOW
        elif "Skipped" in status:  fc = RED
        else:                      fc = LIGHT if i % 2 == 0 else WHITE

        for ci, val in enumerate(row_data, 1):
            c = ws.cell(r, ci, val)
            c.font   = Font(name="Arial", size=10)
            c.fill   = PatternFill("solid", fgColor=fc)
            c.border = bdr
            c.alignment = Alignment(vertical="center", wrap_text=True,
                                     horizontal="center" if ci in (1,) else "left")
            if isinstance(val, str) and val.startswith("http"):
                c.hyperlink = val; c.value = "🔗 View"
                c.font = Font(name="Arial", size=10, color="1155CC", underline="single")
        ws.row_dimensions[r].height = 20

    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{last_col}{len(rows)+2}"


def save_naukri_excel(jobs, job_title, applied_count=0):
    ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(EXCEL_DIR, f"Naukri_{job_title.replace(' ','_')}_{ts}.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Naukri Jobs"
    title_text = f"Naukri Jobs · {job_title.title()} · {datetime.now().strftime('%d %b %Y')} · Applied: {applied_count}"
    headers    = ["#", "Job Title", "Company", "Exp", "Salary", "Location", "Posted", "Link", "Status"]
    rows = []
    for i, job in enumerate(jobs, 1):
        rows.append([
            i, job["title"], job["company"], job.get("exp","N/A"),
            job["salary"], job["location"], job["posted"], job["link"],
            job.get("applied","Pending"),
        ])
    col_widths = [4, 35, 22, 12, 18, 16, 12, 10, 18]
    _excel_style(ws, title_text, headers, rows, col_widths)
    wb.save(filepath)
    print(f"\n💾 Naukri Excel → {filepath}")


def save_linkedin_excel(jobs, job_title, applied_count=0):
    ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(EXCEL_DIR, f"LinkedIn_{job_title.replace(' ','_')}_{ts}.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "LinkedIn Jobs"
    title_text = f"LinkedIn Easy Apply · {job_title.title()} · {datetime.now().strftime('%d %b %Y')} · Applied: {applied_count}"
    headers    = ["#", "Job Title", "Company", "Status", "Time"]
    rows = []
    for i, job in enumerate(jobs, 1):
        rows.append([
            i, job.get("title","N/A"), job.get("company","N/A"),
            job.get("applied","Pending"), job.get("time",""),
        ])
    col_widths = [4, 40, 30, 22, 20]
    _excel_style(ws, title_text, headers, rows, col_widths)
    wb.save(filepath)
    print(f"\n💾 LinkedIn Excel → {filepath}")

# ══════════════════════════════════════════════════════════
# EXECUTE COMMAND
# ══════════════════════════════════════════════════════════

def execute(parsed, browser, page):
    intent    = parsed.get("intent", "unknown")
    job_title = parsed.get("job_title", "Java Developer")
    location  = parsed.get("location", "")
    max_apply = int(parsed.get("max_apply", 5))

    if intent == "naukri_auto_apply":
        if input(f"\n⚠️  Apply to {max_apply} Naukri jobs? (yes/no): ").strip().lower() in ["yes","y"]:
            naukri_auto_apply(browser, page, job_title, location, max_apply)
        else:
            print("Cancelled.")

    elif intent == "naukri_search":
        naukri_search_only(page, job_title, location)

    elif intent == "linkedin_apply":
        if input(f"\n⚠️  Apply to {max_apply} LinkedIn Easy Apply jobs? (yes/no): ").strip().lower() in ["yes","y"]:
            linkedin_auto_apply(browser, page, job_title, location, max_apply)
        else:
            print("Cancelled.")

    elif intent == "linkedin_search":
        linkedin_search(page, job_title, location)

    elif intent == "youtube_search":
        q = parsed.get("query", job_title).replace(" ", "+")
        safe_goto(page, f"https://www.youtube.com/results?search_query={q}")
        print("✅ YouTube opened!")

    elif intent == "google_search":
        q = parsed.get("query", job_title).replace(" ", "+")
        safe_goto(page, f"https://www.google.com/search?q={q}")
        print("✅ Google opened!")

    else:
        print("\n⚠️  Could not understand your command.")
        print("   Examples:")
        print("   → auto apply java developer jobs on naukri")
        print("   → apply to 3 spring boot jobs on naukri in hyderabad")
        print("   → search python developer jobs on naukri")
        print("   → apply to java jobs on linkedin")
        print("   → find java developer jobs on linkedin")

# ══════════════════════════════════════════════════════════
# MAIN  (only used when running ai_agent.py directly)
# When using the dashboard, run server.py instead
# ══════════════════════════════════════════════════════════

def main():
    print("=" * 60)
    print("  🤖 AI PC AGENT v7.1 — FINAL VERSION")
    print("=" * 60)
    print(f"\n  👤 {MY_PROFILE['name']}")
    print(f"  💼 {MY_PROFILE['current_role']} @ {MY_PROFILE['current_company']}")
    print(f"  📧 {MY_PROFILE['email']}  |  📱 {MY_PROFILE['phone']}")
    print(f"\n  💡 TIP: Run server.py for the live dashboard UI")
    print(f"         python server.py → open http://localhost:5000\n")

    if not os.path.exists(MY_PROFILE["resume_path"]):
        print(f"  ⚠️  Resume not found: {MY_PROFILE['resume_path']}\n")

    with sync_playwright() as p:
        print("🚀 Opening browser...")
        browser = p.chromium.launch_persistent_context(
            user_data_dir=AGENT_PROFILE,
            channel="chrome",
            headless=False,
            slow_mo=80,
            args=["--start-maximized", "--disable-blink-features=AutomationControlled"],
            no_viewport=True,
            ignore_default_args=["--enable-automation"],
        )
        page = browser.new_page()
        page.add_init_script("Object.defineProperty(navigator,'webdriver',{get:()=>undefined})")
        print("✅ Browser ready!\n")

        print("─" * 60)
        print("  COMMANDS:")
        print("  • auto apply java developer jobs on naukri")
        print("  • apply to 3 spring boot jobs on naukri in hyderabad")
        print("  • search java jobs on naukri")
        print("  • apply to java jobs on linkedin")
        print("  • find python developer jobs on linkedin")
        print("─" * 60 + "\n")

        while True:
            try:
                command = input("🎤 Command (or 'quit'): ").strip()
            except KeyboardInterrupt:
                print("\n👋 Goodbye!")
                browser.close()
                break

            if command.lower() in ["quit", "exit", "q", "bye"]:
                print("👋 Goodbye!")
                browser.close()
                break
            if not command:
                continue

            parsed = parse_command(command)
            print(f"⚡ intent: {parsed['intent']} | job: '{parsed.get('job_title','')}'"
                  f" | loc: '{parsed.get('location','')}' | max: {parsed.get('max_apply',5)}")

            try:
                execute(parsed, browser, page)
            except Exception as e:
                print(f"\n❌ Error: {e}")
                close_extra_tabs(browser, page)

            print("\n" + "─" * 60 + "\n✅ Ready!\n" + "─" * 60 + "\n")


if __name__ == "__main__":
    main()